import json
import requests
import hashlib
import time
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl
from collections import Counter
from wordcloud import WordCloud
from openpyxl.drawing.image import Image
import os
import traceback

try:
    today_date = datetime.now().strftime('%Y%m%d')

    def generate_sign(params):
        sorted_params = sorted(params.items())
        sign_string = ''.join([f"{key}={value}" for key, value in sorted_params])
        sign_string += "your_secret_key"  # 替换为真实秘钥
        return hashlib.md5(sign_string.encode('utf-8')).hexdigest()

    params = {
        'date': today_date,
        'os': 'android',
        'sv': '8.3.5',
        'ov': '28',
        'net': '',
        'app': 'cailianpress',
        'channel': '6',
        'motif': '0',
        'province_code': '4108',
        'token': '',
        'mb': 'HUAWEI-ELE-AL00',
        'uid': '',
        'sign': '',
        'timestamp': str(int(time.time()))
    }
    params['sign'] = generate_sign(params)

    headers = {
        'accept-encoding': 'gzip',
        'user-agent': 'okhttp/4.9.0'
    }

    url = 'https://x-quote.cls.cn/v2/quote/a/plate/up_down_analysis'
    response = requests.get(url, params=params, headers=headers)

    def save_data_to_excel(plate_stock, continuous_limit_up, file_name='stock_data1.xlsx'):
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        data_font = Font(color="000000")
        data_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        ws_plate_stock = wb.active
        ws_plate_stock.title = "Plate Stock Data"
        headers_plate_stock = ['板块', '涨幅', '上涨原因', '数量', '代码', '名字',
                              '价格', '涨幅', '几天几板', '涨停原因', '概念板块']
        ws_plate_stock.append(headers_plate_stock)
        for col_num, header in enumerate(headers_plate_stock, 1):
            cell = ws_plate_stock.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row_num = 2
        for plate_data in plate_stock:
            change = plate_data.get('change', None)
            change_percentage = change * 100 if change is not None else 'N/A'
            for stock in plate_data.get('stock_list', []):
                stock_change = stock.get('change', None)
                stock_change_percentage = stock_change * 100 if stock_change is not None else 'N/A'
                row = [
                    plate_data.get('secu_name', 'N/A'),
                    change_percentage,
                    plate_data.get('up_reason', 'N/A'),
                    plate_data.get('plate_stock_up_num', 'N/A'),
                    stock.get('secu_code', 'N/A'),
                    stock.get('secu_name', 'N/A'),
                    stock.get('last_px', 'N/A'),
                    stock_change_percentage,
                    stock.get('up_num', 'N/A'),
                    stock.get('up_reason', 'N/A'),
                    ', '.join(stock.get('up_tags', []))
                ]
                for col_num, value in enumerate(row, 1):
                    cell = ws_plate_stock.cell(row=row_num, column=col_num, value=value)
                    cell.font = data_font
                    cell.fill = data_fill
                    cell.alignment = center_alignment
                    cell.border = border
                row_num += 1
        for col in ws_plate_stock.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_plate_stock.column_dimensions[column].width = max_length + 2

        ws_continuous_limit_up = wb.create_sheet(title="Continuous Limit Up")
        headers_continuous_limit_up = ['Height', 'Stock Code', 'Stock Name']
        ws_continuous_limit_up.append(headers_continuous_limit_up)
        for col_num, header in enumerate(headers_continuous_limit_up, 1):
            cell = ws_continuous_limit_up.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row_num = 2
        for item in continuous_limit_up:
            height = item.get('height', 'N/A')
            for stock in item.get('stock_list', []):
                secu_code = stock.get('secu_code', 'N/A')
                secu_name = stock.get('secu_name', 'N/A')
                ws_continuous_limit_up.append([height, secu_code, secu_name])
                for col_num in range(1, 4):
                    cell = ws_continuous_limit_up.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.fill = data_fill
                    cell.alignment = center_alignment
                    cell.border = border
                row_num += 1
        for col in ws_continuous_limit_up.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_continuous_limit_up.column_dimensions[column].width = max_length + 2

        wb.save(file_name)
        print(f"文件已保存到 {file_name}")

    if response.status_code == 200:
        response_data = response.json()
        print("API 响应数据:", response_data)
        if not isinstance(response_data, dict):
            raise ValueError("API 返回的数据不是有效的 JSON 字典")

        plate_stock = response_data.get('data', {}).get('plate_stock', [])
        continuous_limit_up = response_data.get('data', {}).get('continuous_limit_up', [])

        if not plate_stock:
            print("未找到 plate_stock 数据")
        if not continuous_limit_up:
            print("未找到 continuous_limit_up 数据")

        save_data_to_excel(plate_stock, continuous_limit_up, 'stock_data1.xlsx')

        wb = openpyxl.load_workbook('stock_data1.xlsx')
        ws = wb['Plate Stock Data']

        header = [cell.value for cell in ws[1]]
        concept_col_idx = header.index('概念板块') + 1

        all_concepts = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=concept_col_idx, max_col=concept_col_idx):
            cell_value = row[0].value
            if cell_value:
                concepts = [c.strip() for c in cell_value.split(',') if c.strip()]
                all_concepts.extend(concepts)

        counter = Counter(all_concepts)

      
        wordcloud = WordCloud(
            font_path='/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
            background_color='white',
            width=800,
            height=400
        ).generate_from_frequencies(counter)





        
        img_path = 'wordcloud.png'
        wordcloud.to_file(img_path)

        if '词云图' in wb.sheetnames:
            del wb['词云图']
        ws_img = wb.create_sheet('词云图')

        img = Image(img_path)
        ws_img.add_image(img, 'A1')

        wb.save('stock_data1.xlsx')
        print(f"词云图片已成功插入到 stock_data1.xlsx 的新工作表 '词云图' 中")

    else:
        print(f"API 请求失败，状态码: {response.status_code}")

except Exception:
    print("执行异常，详细信息：")
    traceback.print_exc()
    exit(1)
